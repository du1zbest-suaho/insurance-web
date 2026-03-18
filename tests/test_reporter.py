"""
test_reporter.py — xlsx 생성 및 미리보기 빌드 테스트

검증 항목:
  - build_product_mapping: upper = dtcd+itcd (2061001)
  - build_product_mapping: __default__ fallback
  - build_product_mapping: upper_name / lower_name 설정
  - generate_xlsx: 실제 템플릿 사용 시 4행 헤더 기반 기록
  - generate_xlsx: 템플릿 없을 때 기본 헤더 자동 생성
  - generate_xlsx: SET_CODE 필드 포함 (핵심 수정 검증)
  - generate_xlsx: UPPER_OBJECT_NAME / LOWER_OBJECT_NAME 필드 포함 (핵심 수정 검증)
  - generate_xlsx: VALID_START_DATE 기록
  - generate_xlsx: 7행부터 데이터 시작
  - build_preview: GT 비교 결과 포함 (annotated_rows)
  - build_preview: missing_rows 포함
"""

import os

import openpyxl
import pytest

from app.core.reporter import build_preview, build_product_mapping, generate_xlsx
from app.core.converter import TABLE_CONVERTERS


# ─── 헬퍼 ─────────────────────────────────────────────────────────────────────

def read_xlsx_row(path: str, row_num: int) -> dict:
    """xlsx 4행 헤더 기준으로 특정 행을 dict로 반환"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = {str(cell.value).strip(): cell.column
               for cell in ws[4] if cell.value}
    return {
        name: ws.cell(row=row_num, column=col).value
        for name, col in headers.items()
    }


# ─── 1. build_product_mapping ────────────────────────────────────────────────

class TestBuildProductMapping:
    def test_upper_is_dtcd_plus_itcd(self, synthetic_mapping_entries):
        """upper_object_code = dtcd + itcd (예: '2061001')"""
        pm = build_product_mapping(synthetic_mapping_entries)
        assert pm["__default__"]["upper"] == "2061001"

    def test_default_key_exists(self, synthetic_mapping_entries):
        pm = build_product_mapping(synthetic_mapping_entries)
        assert "__default__" in pm

    def test_upper_name_from_sale_nm(self, synthetic_mapping_entries):
        pm = build_product_mapping(synthetic_mapping_entries)
        assert pm["__default__"]["upper_name"] == "테스트보험(표준체형)"

    def test_lower_name_from_prod_sale_nm(self, synthetic_mapping_entries):
        pm = build_product_mapping(synthetic_mapping_entries)
        assert pm["__default__"]["lower_name"] == "테스트보험"

    def test_empty_entries_returns_empty(self):
        pm = build_product_mapping([])
        assert pm == {}

    def test_multiple_entries_first_is_default(self):
        entries = [
            {"dtcd": "1000", "itcd": "001", "sale_nm": "A보험",
             "prod_dtcd": "1000", "prod_itcd": "1", "prod_sale_nm": "A상품"},
            {"dtcd": "2000", "itcd": "002", "sale_nm": "B보험",
             "prod_dtcd": "2000", "prod_itcd": "2", "prod_sale_nm": "B상품"},
        ]
        pm = build_product_mapping(entries)
        assert pm["__default__"]["upper"] == "1000001"

    def test_sub_type_fallback_uses_default(self, synthetic_mapping_entries):
        pm = build_product_mapping(synthetic_mapping_entries)
        # 존재하지 않는 sub_type → __default__ fallback
        mapping = pm.get("존재하지않는종목") or pm.get("__default__", {})
        assert mapping["upper"] == "2061001"

    def test_lower_object_code_zero_padded(self, synthetic_mapping_entries):
        """B7: prod_itcd='1', itcd='001'(len=3) → lower='2061001' (20611 아님)"""
        pm = build_product_mapping(synthetic_mapping_entries)
        assert pm["__default__"]["lower"] == "2061001", (
            f"lower={pm['__default__']['lower']} — zero-pad 버그: '20611' 이면 수정 필요"
        )

    def test_lower_object_code_nondigit_prod_itcd_not_padded(self):
        """prod_itcd가 숫자가 아닌 경우 그대로 사용"""
        entries = [
            {"dtcd": "2257", "itcd": "A01", "sale_nm": "시그니처",
             "prod_dtcd": "2257", "prod_itcd": "A01", "prod_sale_nm": "시그니처보험"},
        ]
        pm = build_product_mapping(entries)
        assert pm["__default__"]["lower"] == "2257A01"

    def test_lower_equals_upper_when_no_prod_info(self):
        """prod_dtcd/prod_itcd 없으면 lower = upper"""
        entries = [
            {"dtcd": "2061", "itcd": "001", "sale_nm": "테스트",
             "prod_dtcd": "", "prod_itcd": "", "prod_sale_nm": ""},
        ]
        pm = build_product_mapping(entries)
        assert pm["__default__"]["lower"] == pm["__default__"]["upper"]


# ─── 2. generate_xlsx — 템플릿 없음 ──────────────────────────────────────────

class TestGenerateXlsxNoTemplate:
    def test_creates_file(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "out_S00026.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        assert os.path.exists(out)

    def test_header_row_4_has_columns(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "hdr.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[4] if c.value]
        assert "UPPER_OBJECT_CODE" in headers
        assert "SET_CODE" in headers

    def test_set_code_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """SET_CODE 필드가 xlsx에 기록되어야 한다 (핵심 수정)"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "set_code.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("SET_CODE") == "S00026", f"SET_CODE={row7.get('SET_CODE')}"

    def test_upper_object_code_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """UPPER_OBJECT_CODE = dtcd+itcd (핵심 수정)"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "upper_code.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("UPPER_OBJECT_CODE") == "2061001", (
            f"UPPER_OBJECT_CODE={row7.get('UPPER_OBJECT_CODE')}"
        )

    def test_upper_object_name_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """UPPER_OBJECT_NAME 필드가 기록되어야 한다 (핵심 수정)"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "upper_name.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("UPPER_OBJECT_NAME") == "테스트보험(표준체형)", (
            f"UPPER_OBJECT_NAME={row7.get('UPPER_OBJECT_NAME')}"
        )

    def test_valid_start_date_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "date.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("VALID_START_DATE") == "20260101"

    def test_data_starts_at_row_7(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """데이터가 7행부터 시작해야 한다"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "row7.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        # 6행은 비어 있어야 함 (헤더 영역)
        row6_vals = [c.value for c in ws[6] if c.value]
        row7_vals = [c.value for c in ws[7] if c.value]
        assert len(row6_vals) == 0, f"6행에 데이터가 있음: {row6_vals}"
        assert len(row7_vals) > 0, "7행에 데이터가 없음"

    def test_lower_object_code_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """B7: LOWER_OBJECT_CODE = prod_dtcd + prod_itcd.zfill(len(itcd)) = '2061001'"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "lower_code.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("LOWER_OBJECT_CODE") == "2061001", (
            f"LOWER_OBJECT_CODE={row7.get('LOWER_OBJECT_CODE')} — zero-pad 수정 검증"
        )

    def test_sale_chnl_code_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """SALE_CHNL_CODE = '1,2,3,4,7' (고정값)"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "chnl.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("SALE_CHNL_CODE") == "1,2,3,4,7"

    def test_valid_end_date_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        """VALID_END_DATE = '99991231' (고정값)"""
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "enddate.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("VALID_END_DATE") == "99991231"

    def test_all_coded_rows_written(self, tmp_path, synthetic_mapping_entries, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "all_rows.xlsx")
        generate_xlsx("S00026", coded, None, out, synthetic_mapping_entries)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        data_rows = [r for r in ws.iter_rows(min_row=7, values_only=True)
                     if any(v is not None for v in r)]
        assert len(data_rows) == len(coded), (
            f"기록된 행 수({len(data_rows)}) != coded_rows 수({len(coded)})"
        )


# ─── 3. generate_xlsx — 실제 템플릿 사용 ────────────────────────────────────

class TestGenerateXlsxWithTemplate:
    def test_creates_file_with_template(
        self, tmp_path, template_paths, synthetic_mapping_entries, raw_rows_s00026
    ):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "with_template.xlsx")
        generate_xlsx("S00026", coded, template_paths["S00026"], out,
                      synthetic_mapping_entries, "20260101")
        assert os.path.exists(out)

    def test_template_set_code_written(
        self, tmp_path, template_paths, synthetic_mapping_entries, raw_rows_s00026
    ):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "tmpl_setcode.xlsx")
        generate_xlsx("S00026", coded, template_paths["S00026"], out,
                      synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("SET_CODE") == "S00026"

    def test_template_upper_object_code(
        self, tmp_path, template_paths, synthetic_mapping_entries, raw_rows_s00026
    ):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        out = str(tmp_path / "tmpl_upper.xlsx")
        generate_xlsx("S00026", coded, template_paths["S00026"], out,
                      synthetic_mapping_entries, "20260101")
        row7 = read_xlsx_row(out, 7)
        assert row7.get("UPPER_OBJECT_CODE") == "2061001"

    def test_s00022_fpin_spin_fields_written(
        self, tmp_path, template_paths, synthetic_mapping_entries, raw_rows_s00022
    ):
        """B8: S00022 xlsx에 FPIN/SPIN_STRT_DVSN_VAL 기록 (MIN_AG/MAX_AG 아님)"""
        coded = TABLE_CONVERTERS["S00022"](raw_rows_s00022)
        out = str(tmp_path / "s00022_fpin.xlsx")
        generate_xlsx("S00022", coded, template_paths["S00022"], out,
                      synthetic_mapping_entries, "20260101")
        assert os.path.exists(out)
        wb = openpyxl.load_workbook(out, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in ws[4] if c.value]
        # 템플릿에 FPIN/SPIN 컬럼이 있으면 데이터가 기록되어야 함
        if "FPIN_STRT_DVSN_VAL" in headers:
            row7 = read_xlsx_row(out, 7)
            assert row7.get("FPIN_STRT_DVSN_VAL") is not None, (
                "FPIN_STRT_DVSN_VAL가 기록되지 않음 — S00022 COLUMN_MAPPINGS 수정 검증"
            )

    def test_s00022_no_template_has_correct_headers(
        self, tmp_path, synthetic_mapping_entries, raw_rows_s00022
    ):
        """B8: S00022 템플릿 없는 경우 FPIN/SPIN 헤더 포함 (MIN_AG/MAX_AG 아님)"""
        coded = TABLE_CONVERTERS["S00022"](raw_rows_s00022)
        out = str(tmp_path / "s00022_noltempl.xlsx")
        generate_xlsx("S00022", coded, None, out, synthetic_mapping_entries)
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [c.value for c in ws[4] if c.value]
        assert "FPIN_STRT_DVSN_VAL" in headers, (
            f"FPIN_STRT_DVSN_VAL가 헤더에 없음. 헤더={headers}"
        )
        assert "MIN_AG" not in headers, "MIN_AG는 S00022 헤더에 있으면 안됨"

    @pytest.mark.parametrize("tbl", ["S00027", "S00028", "S00022"])
    def test_other_tables_with_template(
        self, tmp_path, template_paths, synthetic_mapping_entries,
        raw_rows_s00027, raw_rows_s00028, raw_rows_s00022, tbl
    ):
        raw_map = {
            "S00027": raw_rows_s00027,
            "S00028": raw_rows_s00028,
            "S00022": raw_rows_s00022,
        }
        coded = TABLE_CONVERTERS[tbl](raw_map[tbl])
        out = str(tmp_path / f"tmpl_{tbl}.xlsx")
        generate_xlsx(tbl, coded, template_paths[tbl], out,
                      synthetic_mapping_entries, "20260101")
        assert os.path.exists(out)
        row7 = read_xlsx_row(out, 7)
        assert row7.get("SET_CODE") == tbl


# ─── 4. build_preview ────────────────────────────────────────────────────────

class TestBuildPreview:
    def test_basic_preview(self, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        preview = build_preview("S00026", coded, "샘플 텍스트")
        assert preview["table_type"] == "S00026"
        assert preview["count"] == len(coded)
        assert len(preview["headers"]) > 0
        assert len(preview["rows"]) == len(coded)

    def test_preview_with_gt_match(self, raw_rows_s00026, synthetic_gt_rows_s00026):
        from app.core.comparator import compare
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        gt = [r for r in synthetic_gt_rows_s00026 if r["MAX_AG"] != 999]
        gt_comp = compare("S00026", coded, gt)
        preview = build_preview("S00026", coded, "", gt_comparison=gt_comp)
        assert preview["has_gt"] is True
        statuses = {r["_gt_status"] for r in preview["rows"]}
        assert statuses.issubset({"match", "new", "missing"})

    def test_preview_missing_rows_appended(self, raw_rows_s00026, synthetic_gt_rows_s00026):
        from app.core.comparator import compare
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026[:1])  # 일부만
        gt = [r for r in synthetic_gt_rows_s00026 if r["MAX_AG"] != 999]
        gt_comp = compare("S00026", coded, gt)
        preview = build_preview("S00026", coded, "", gt_comparison=gt_comp)
        missing = [r for r in preview["rows"] if r["_gt_status"] == "missing"]
        assert len(missing) == preview["missing_count"]
        assert len(missing) > 0

    def test_preview_no_gt(self, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        preview = build_preview("S00026", coded)
        assert preview["has_gt"] is False
        for r in preview["rows"]:
            assert r["_gt_status"] is None

    def test_preview_text_snippet_truncated(self, raw_rows_s00026):
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        long_text = "A" * 5000
        preview = build_preview("S00026", coded, long_text)
        assert len(preview["text_snippet"]) <= 3000

    def test_preview_compare_fields_included(self, raw_rows_s00026, synthetic_gt_rows_s00026):
        from app.core.comparator import compare
        coded = TABLE_CONVERTERS["S00026"](raw_rows_s00026)
        gt = [r for r in synthetic_gt_rows_s00026 if r["MAX_AG"] != 999]
        gt_comp = compare("S00026", coded, gt)
        preview = build_preview("S00026", coded, "", gt_comparison=gt_comp)
        assert "compare_fields" in preview
        assert len(preview["compare_fields"]) > 0
