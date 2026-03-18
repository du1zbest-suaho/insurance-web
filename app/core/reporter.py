"""
reporter.py — 미리보기 데이터 빌더 + xlsx 생성
ruleautomatker/.claude/skills/xlsx-generator/scripts/generate_upload.py 로직 재사용

핵심 변경:
- build_product_mapping: sub_type → object codes, __default__ fallback, upper = dtcd+itcd
- generate_xlsx: SET_CODE / UPPER_OBJECT_NAME / LOWER_OBJECT_NAME 필드 추가
- GT 비교 결과(annotated_rows + missing_rows) 포함
"""

import os
import shutil
import sys
from typing import Any, Dict, List, Optional

# ruleautomatker의 generate_upload.py import
_RULE_BASE = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "../../../ruleautomatker")
)
_XLSX_DIR = os.path.join(_RULE_BASE, ".claude/skills/xlsx-generator/scripts")

if _XLSX_DIR not in sys.path:
    sys.path.insert(0, _XLSX_DIR)

try:
    import openpyxl
    from generate_upload import DEFAULT_SALE_CHNL_CODE
    from generate_upload import COLUMN_MAPPINGS as _BASE_COLUMN_MAPPINGS
except ImportError:
    _BASE_COLUMN_MAPPINGS = {}
    DEFAULT_SALE_CHNL_CODE = "1,2,3,4,7"

# S00022: generate_upload.py의 MIN_AG/MAX_AG 매핑은 오류.
# 실제 템플릿 컬럼(FPIN_STRT_AG_INQY_CODE 등)과 converter 출력 필드를 직접 매핑.
COLUMN_MAPPINGS = {
    **_BASE_COLUMN_MAPPINGS,
    "S00022": {
        "FPIN_STRT_AG_INQY_CODE": "FPIN_STRT_AG_INQY_CODE",
        "SPIN_STRT_AG_INQY_CODE": "SPIN_STRT_AG_INQY_CODE",
        "FPIN_STRT_DVSN_CODE":    "FPIN_STRT_DVSN_CODE",
        "FPIN_STRT_DVSN_VAL":     "FPIN_STRT_DVSN_VAL",
        "SPIN_STRT_DVSN_CODE":    "SPIN_STRT_DVSN_CODE",
        "SPIN_STRT_DVSN_VAL":     "SPIN_STRT_DVSN_VAL",
    },
}


def build_product_mapping(mapping_entries: List[Dict]) -> Dict[str, Dict]:
    """
    mapping_entries → sub_type 기반 object code 매핑.

    ruleautomatker generate_upload.py 동일 로직:
    - sub_type을 키로 사용
    - upper_object_code = dtcd + itcd (예: "2061" + "001" = "2061001")
    - 첫 번째 항목을 __default__로 설정 (sub_type 불일치 시 fallback)
    - 여러 엔트리는 각각 저장 (itcd를 fallback 키로도 추가)
    """
    if not mapping_entries:
        return {}

    pm: Dict[str, Dict] = {}

    for entry in mapping_entries:
        dtcd = str(entry.get("dtcd", "")).strip()
        itcd = str(entry.get("itcd", "")).strip()
        sale_nm = str(entry.get("sale_nm", "")).strip()
        prod_dtcd = str(entry.get("prod_dtcd", "")).strip()
        prod_itcd = str(entry.get("prod_itcd", "")).strip()
        prod_sale_nm = str(entry.get("prod_sale_nm", "")).strip()

        # upper_object_code = dtcd + itcd (ruleautomatker 표준, itcd는 이미 포맷됨)
        upper = (dtcd + itcd) if dtcd and itcd else dtcd or itcd

        # lower_object_code = prod_dtcd + prod_itcd
        # prod_itcd는 매핑 xlsx에서 숫자 단독("1")으로 올 수 있음.
        # itcd와 동일 길이로 zero-pad하여 코드 형식을 맞춤.
        # 예: itcd="001"(3자리), prod_itcd="1" → "001" → lower="2061001"
        #     itcd="A01"(3자리), prod_itcd="1" → "001" → lower="2257001"
        if prod_dtcd and prod_itcd and itcd:
            target_len = len(itcd)
            # 숫자로만 구성된 prod_itcd는 zero-pad, 아니면 그대로 사용
            padded = prod_itcd.zfill(target_len) if prod_itcd.isdigit() else prod_itcd
            lower = prod_dtcd + padded
        elif prod_dtcd and prod_itcd:
            lower = prod_dtcd + prod_itcd
        else:
            lower = upper

        entry_map = {
            "upper": upper,
            "upper_name": sale_nm,
            "lower": lower,
            "lower_name": prod_sale_nm or sale_nm,
        }

        # 첫 번째 항목 → __default__
        if "__default__" not in pm:
            pm["__default__"] = entry_map

        # itcd를 sub_type 후보 키로도 등록 (extraction sub_type과 일치 가능성)
        if itcd and itcd not in pm:
            pm[itcd] = entry_map

    return pm


def generate_xlsx(
    table_type: str,
    coded_rows: List[Dict],
    template_path: Optional[str],
    output_path: str,
    mapping_entries: List[Dict],
    valid_start_date: str = "",
    valid_end_date: str = "99991231",
) -> str:
    """
    xlsx 업로드 양식 생성. 반환: output_path

    ruleautomatker generate_upload.py 동일 로직:
    - 4행 헤더 기준 컬럼 인덱스
    - 7행부터 데이터
    - SET_CODE / UPPER_OBJECT_NAME / LOWER_OBJECT_NAME 포함
    - sub_type → __default__ fallback
    """
    os.makedirs(
        os.path.dirname(output_path) if os.path.dirname(output_path) else ".",
        exist_ok=True,
    )
    product_mapping = build_product_mapping(mapping_entries)

    HEADER_ROW = 4
    DATA_START_ROW = 7

    if template_path and os.path.exists(template_path):
        shutil.copy2(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # 기존 예시 데이터(7행~) 삭제
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        # 템플릿 없음 → 공통 + 테이블별 컬럼을 4행에 직접 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        col_map = COLUMN_MAPPINGS.get(table_type, {})
        default_headers = [
            "UPPER_OBJECT_CODE", "UPPER_OBJECT_NAME",
            "LOWER_OBJECT_CODE", "LOWER_OBJECT_NAME",
            "SET_CODE", "VALID_START_DATE", "VALID_END_DATE", "SALE_CHNL_CODE",
        ] + list(col_map.keys())
        for col_idx, name in enumerate(default_headers, start=1):
            ws.cell(row=HEADER_ROW, column=col_idx, value=name)

    # 4행에서 컬럼 헤더 위치 파악
    col_index: Dict[str, int] = {}
    for cell in ws[HEADER_ROW]:
        if cell.value:
            col_index[str(cell.value).strip()] = cell.column

    def set_cell(row_idx: int, col_name: str, value: Any) -> None:
        if col_name in col_index:
            ws.cell(row=row_idx, column=col_index[col_name], value=value)

    col_map = COLUMN_MAPPINGS.get(table_type, {})

    for row_offset, coded_row in enumerate(coded_rows):
        excel_row = DATA_START_ROW + row_offset
        sub_type = str(coded_row.get("sub_type", "")).strip()

        # sub_type 매칭 → fallback __default__ (ruleautomatker 동일)
        mapping = product_mapping.get(sub_type) or product_mapping.get("__default__", {})

        # _upper/lower_object_code 우선 (특약 등 미리 세팅된 경우)
        upper      = coded_row.get("_upper_object_code") or mapping.get("upper", "")
        upper_name = coded_row.get("_upper_object_name") or mapping.get("upper_name", "")
        lower      = coded_row.get("_lower_object_code") or mapping.get("lower", "")
        lower_name = coded_row.get("_lower_object_name") or mapping.get("lower_name", "")

        set_cell(excel_row, "UPPER_OBJECT_CODE", upper)
        set_cell(excel_row, "UPPER_OBJECT_NAME", upper_name)
        set_cell(excel_row, "LOWER_OBJECT_CODE", lower)
        set_cell(excel_row, "LOWER_OBJECT_NAME", lower_name)
        set_cell(excel_row, "SET_CODE", table_type)
        set_cell(excel_row, "VALID_START_DATE", valid_start_date)
        set_cell(excel_row, "VALID_END_DATE", valid_end_date)
        set_cell(excel_row, "SALE_CHNL_CODE", DEFAULT_SALE_CHNL_CODE)

        # 테이블 타입별 데이터 컬럼
        for xlsx_col, data_key in col_map.items():
            val = coded_row.get(data_key)
            if val is not None:
                set_cell(excel_row, xlsx_col, val)

    wb.save(output_path)
    return output_path


def build_preview(
    table_type: str,
    coded_rows: List[Dict],
    raw_text_snippet: str = "",
    gt_comparison: Optional[Dict] = None,
) -> Dict[str, Any]:
    """
    미리보기 JSON 빌드 (프론트엔드용).
    gt_comparison이 있으면 annotated_rows + missing_rows 포함.
    각 행에 _gt_status: "match" | "new" | "missing" | None 추가.
    """
    col_map = COLUMN_MAPPINGS.get(table_type, {})
    headers = list(col_map.keys()) if col_map else ["sub_type"]

    has_gt = gt_comparison is not None and gt_comparison.get("has_gt", False)
    source_rows = gt_comparison["annotated_rows"] if has_gt else coded_rows
    missing_rows = gt_comparison.get("missing_rows", []) if has_gt else []
    gt_summary = gt_comparison.get("summary") if gt_comparison else None
    compare_fields = gt_comparison.get("compare_fields", []) if gt_comparison else []

    def _build_row(r: Dict, force_gt_status: Optional[str] = None) -> Dict:
        row_data: Dict[str, Any] = {"sub_type": str(r.get("sub_type", "") or "")}
        for xlsx_col, data_key in col_map.items():
            row_data[xlsx_col] = r.get(data_key)
        row_data["_warnings"] = r.get("_warnings", [])
        row_data["_convert_error"] = str(r.get("_convert_error", "") or "")
        row_data["_gt_status"] = force_gt_status or r.get("_gt_status")
        return row_data

    rows = [_build_row(r) for r in source_rows]
    rows += [_build_row(r, "missing") for r in missing_rows]

    return {
        "table_type": table_type,
        "headers": headers,
        "rows": rows,
        "count": len(source_rows),
        "missing_count": len(missing_rows),
        "has_gt": has_gt,
        "gt_summary": gt_summary,
        "compare_fields": compare_fields,
        "text_snippet": raw_text_snippet[:3000] if raw_text_snippet else "",
    }
